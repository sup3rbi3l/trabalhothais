
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv





df_Grupo = load_workbook('Grupo de indicação Outubro 2023.xlsx')
dados_cidades = {
    "Alegrete": '390',
    "Almirante Tamandaré": '294',
    "Alvorada": '458',
    "Apucarana": '322',
    "Arapongas": '022',
    "Araranguá": '400',
    "Araucária": '333',
    "Arroio do Meio": '618',
    "Bagé": '690',
    "Balneário Camboriú": '752',
    "Bento Gonçalves": '688',
    "Biguaçu": '058',
    "Blumenau": '700',
    "Brusque": '070',
    "Caçador": '819',
    "Cachoeira do Sul": '394',
    "Cachoeirinha": '560',
    "Camaquã": '466',
    "Cambé": '379',
    "Camboriú": '868',
    "Campo Bom": '584',
    "Campo Largo": '389',
    "Canela": '601',
    "Canoas": '603',
    "Capão da Canoa": '87',
    "Carazinho": '614',
    "Carlos Barbosa": '920',
    "Cascavel": '641',
    "Caxias do Sul": '687',
    "Chapecó": '068',
    "Charqueadas": '217',
    "Cianorte": '430',
    "Colombo": '437',
    "Concórdia": '113',
    "Criciúma": '89',
    "Cruz Alta": '685',
    "Curitiba": '884',
    "Dois Irmãos": '781',
    "Eldorado do Sul": '918',
    "Encantado": '934',
    "Erechim": '695',
    "Estância Velha": '757',
    "Esteio": '758',
    "Estrela": '231',
    "Farroupilha": '83',
    "Florianópolis": '88',
    "Foz do Iguaçu": '506',
    "Fraiburgo": '449',
    "Frederico Westphalen": '536',
    "Garibaldi": '568',
    "Gaspar": '154',
    "Gramado": '817',
    "Gravataí": '720',
    "Guaíba": '728',
    "Guaramirim": '163',
    "Guarapuava": '531',
    "Herval d'Oeste": '698',
    "Ibiporã": '204',
    "Içara": '724',
    "Igrejinha": '885',
    "Imbé": '835',
    "Indaial": '775',
    "Itajaí": '193',
    "Itapema": '94',
    "Itaqui": '106',
    "Ivoti": '173',
    "Jaraguá do Sul": '203',
    "Joaçaba": '119',
    "Joinville": '086',
    "Lages": '213',
    "Lajeado": '689',
    "Londrina": '996',
    "Mafra": '348',
    "Marau": '638',
    "Maringá": '091',
    "Montenegro": '950',
    "Navegantes": '260',
    "Nova Petrópolis": '201',
    "Novo Hamburgo": '686',
    "Osório": '293',
    "Palhoça": '278',
    "Palmas": '924',
    "Palmeira das Missões": '412',
    "Panambi": '447',
    "Paranaguá": '749',
    "Parobé": '499',
    "Passo Fundo": '693',
    "Pelotas": '691',
    "Pinhais": '770',
    "Piraquara": '380',
    "Ponta Grossa": '794',
    "Porto Alegre": '78',
    "Rio do Sul": '301',
    "Rio Grande": '692',
    "Rio Negrinho": '352',
    "Rio Pardo": '478',
    "Rolândia": '851',
    "Rosário do Sul": '613',
    "Sant'Ana do Livramento": '041',
    "Santa Cruz do Sul": '694',
    "Santa Maria": '075',
    "Santa Rosa": '940',
    "Santo Ângelo": '77',
    "São Bento do Sul": '570',
    "São Borja": '229',
    "São Francisco do Sul": '620',
    "São Gabriel": '270',
    "São José": '363',
    "São José dos Pinhais": '923',
    "São Leopoldo": '710',
    "São Lourenço do Sul": '482',
    "São Luiz Gonzaga": '504',
    "Sapiranga": '330',
    "Sapucaia do Sul": '332',
    "Taquara": '242',
    "Telêmaco Borba": '331',
    "Teutônia": '324',
    "Timbó": '026',
    "Torres": '398',
    "Tramandaí": '438',
    "Três Coroas": '497',
    "Tubarão": '090',
    "União da Vitória": '593',
    "Uruguaiana": '684',
    "Vacaria": '764',
    "Venâncio Aires": '829',
    "Vera Cruz": '839',
    "Veranópolis": '853',
    "Viamão": '871',
    "Videira": '236',
    "Xanxerê": '334',
    "Xaxim": '361',
}




P2 = Workbook()
P2_ativo = P2.active
with open('P2 Outubro 2023.csv') as f:
    reader = csv.reader(f,delimiter=";")
    for row in reader:
        P2_ativo.append(row)
        
df_Grupo = load_workbook('Grupo de indicação Outubro 2023.xlsx')
grupo1 = df_Grupo.worksheets[0]
                    #1 estranho

for row in grupo1:
        if row[16].value == 'Cidade' or row[12].value == '---':
            pass
        else:
            if row[16].value == None:
                break
            
            contrato_completo = str (dados_cidades [row[16].value]+str(row[12].value))
            
            #print(contrato_completo)
            contrato_encontrado = True
            for rows in P2_ativo:
                
                
                if rows[19].value == contrato_completo:
                    
                    if rows[2].value[0:7] == 'ESTORNO':
                        print('a')
                    
                    print('/valor franquia',rows[12].value,contrato_completo)
                    a = False
                    break
            if contrato_encontrado:
                print (contrato_completo)
            
