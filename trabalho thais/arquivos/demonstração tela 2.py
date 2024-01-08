
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from kivy.lang import Builder

from kivymd.app import MDApp

from kivymd.uix.screen import MDScreen
from kivymd.app import MDApp
from kivy.uix.image import Image
from kivymd.uix.button import MDFillRoundFlatIconButton, MDFillRoundFlatButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivymd.uix.toolbar import MDToolbar
from kivy.config import Config
from kivy.app import App
from kivy import platform
from kivymd.uix.filemanager import MDFileManager 
from kivymd.uix.selectioncontrol import MDCheckbox
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from datetime import date


screen = '''

ScreenManager:
    Tela:
    
<Tela>:
    name:'tela'
    
    MDToolbar:
        title: "Buscador"
        pos_hint: {"top" : 1}

    MDLabel:
        font_style:'H6'
        text:'Linha:'
        pos_hint: {"center_x":0.65,"center_y":0.7}
        
    MDLabel:
        font_style:'H6'
        id:linha
        text:''
        pos_hint: {"center_x":0.65,"center_y":0.6}
        
    MDLabel:
        font_style:'H6'
        text:'Data:'
        pos_hint: {"center_x":0.85,"center_y":0.7}
    MDLabel:
        font_style:'H6'
        id:data
        text:''
        pos_hint: {"center_x":0.85,"center_y":0.6}
    
    MDLabel:
        font_style:'H6'
        text:'Comição:'
        pos_hint: {"center_x":1.05,"center_y":0.7}
    MDLabel:
        font_style:'H6'
        id:comicao
        text:''
        pos_hint: {"center_x":1.05,"center_y":0.6}
        
    MDLabel:
        font_style:'H6'
        text:'Franquia:'
        pos_hint: {"center_x":1.25,"center_y":0.7}
    MDLabel:
        font_style:'H6'
        id:franquia
        text:''
        pos_hint: {"center_x":1.25,"center_y":0.6}

    
    MDCheckbox:
        group:'1'
        size_hint: None, None
        size: "48dp", "48dp"
        pos_hint: {'center_x': .9, 'center_y': .25}
        on_active: app.P1()
        
    MDCheckbox:
        group:'1'
        size_hint: None, None
        size: "48dp", "48dp"
        pos_hint: {'center_x': .9, 'center_y': .15}
        on_active: app.P2()

    MDLabel:
        text:'P1'
        font_style:'H6'
        pos_hint: {'center_x': 1.3, 'center_y': .25}
        
    MDLabel:
        text:'P2'
        font_style:'H6'
        pos_hint: {'center_x': 1.3, 'center_y': .15}
    
    
    MDTextField:
        id: contrato
        size_hint_x:(0.4)
        font_size : 22
        pos_hint : {"center_x":0.5,"center_y":0.4}
        hint_text: "contrato"
    
    MDFillRoundFlatButton:
        text:'buscar'
        font_size:20
        pos_hint: {"center_x":0.5,"center_y":0.3,}
        size_hint:(0.1,0.1)
        on_press: app.mostraInfo()

        
'''

class Tela(Screen):
    pass


sm = ScreenManager()
sm.add_widget(Tela(name='tela'))

class procuraContrato(MDApp):
    def build(self):
        self.screen=Builder.load_string(screen)
        self.dados_cidades = {
            
    "Alegrete": '390',
    "Almirante Tamandaré": '294',
    "Alvorada": '458',
    "Apucarana": '322',
    "Arapongas": '022',
    "Araranguá": '400',
    "Araucária": '333',
    "Arroio do Meio": '618',
    "Bagé": '690',
    "Balneário Camboriú": '752',
    "Bento Gonçalves": '688',
    "Biguaçu": '058',
    "Blumenau": '700',
    "Brusque": '070',
    "Caçador": '819',
    "Cachoeira do Sul": '394',
    "Cachoeirinha": '560',
    "Camaquã": '466',
    "Cambé": '379',
    "Camboriú": '868',
    "Campo Bom": '584',
    "Campo Largo": '389',
    "Canela": '601',
    "Canoas": '603',
    "Capão da Canoa": '87',
    "Carazinho": '614',
    "Carlos Barbosa": '920',
    "Cascavel": '641',
    "Caxias do Sul": '687',
    "Chapecó": '068',
    "Charqueadas": '217',
    "Cianorte": '430',
    "Colombo": '437',
    "Concórdia": '113',
    "Criciúma": '89',
    "Cruz Alta": '685',
    "Curitiba": '884',
    "Dois Irmãos": '781',
    "Eldorado do Sul": '918',
    "Encantado": '934',
    "Erechim": '695',
    "Estância Velha": '757',
    "Esteio": '758',
    "Estrela": '231',
    "Farroupilha": '83',
    "Florianópolis": '88',
    "Foz do Iguaçu": '506',
    "Fraiburgo": '449',
    "Frederico Westphalen": '536',
    "Garibaldi": '568',
    "Gaspar": '154',
    "Gramado": '817',
    "Gravataí": '720',
    "Guaíba": '728',
    "Guaramirim": '163',
    "Guarapuava": '531',
    "Herval d'Oeste": '698',
    "Ibiporã": '204',
    "Içara": '724',
    "Igrejinha": '885',
    "Imbé": '835',
    "Indaial": '775',
    "Itajaí": '193',
    "Itapema": '94',
    "Itaqui": '106',
    "Ivoti": '173',
    "Jaraguá do Sul": '203',
    "Joaçaba": '119',
    "Joinville": '086',
    "Lages": '213',
    "Lajeado": '689',
    "Londrina": '996',
    "Mafra": '348',
    "Marau": '638',
    "Maringá": '091',
    "Montenegro": '950',
    "Navegantes": '260',
    "Nova Petrópolis": '201',
    "Novo Hamburgo": '686',
    "Osório": '293',
    "Palhoça": '278',
    "Palmas": '924',
    "Palmeira das Missões": '412',
    "Panambi": '447',
    "Paranaguá": '749',
    "Parobé": '499',
    "Passo Fundo": '693',
    "Pelotas": '691',
    "Pinhais": '770',
    "Piraquara": '380',
    "Ponta Grossa": '794',
    "Porto Alegre": '78',
    "Rio do Sul": '301',
    "Rio Grande": '692',
    "Rio Negrinho": '352',
    "Rio Pardo": '478',
    "Rolândia": '851',
    "Rosário do Sul": '613',
    "Sant'Ana do Livramento": '041',
    "Santa Cruz do Sul": '694',
    "Santa Maria": '075',
    "Santa Rosa": '940',
    "Santo Ângelo": '77',
    "São Bento do Sul": '570',
    "São Borja": '229',
    "São Francisco do Sul": '620',
    "São Gabriel": '270',
    "São José": '363',
    "São José dos Pinhais": '923',
    "São Leopoldo": '710',
    "São Lourenço do Sul": '482',
    "São Luiz Gonzaga": '504',
    "Sapiranga": '330',
    "Sapucaia do Sul": '332',
    "Taquara": '242',
    "Telêmaco Borba": '331',
    "Teutônia": '324',
    "Timbó": '026',
    "Torres": '398',
    "Tramandaí": '438',
    "Três Coroas": '497',
    "Tubarão": '090',
    "União da Vitória": '593',
    "Uruguaiana": '684',
    "Vacaria": '764',
    "Venâncio Aires": '829',
    "Vera Cruz": '839',
    "Veranópolis": '853',
    "Viamão": '871',
    "Videira": '236',
    "Xanxerê": '334',
    "Xaxim": '361',
}
        
        P1_df = Workbook()
        self.P1_ativo = P1_df.active
        with open('P1 Outubro 2023.csv') as f:
            reader = csv.reader(f,delimiter=";")
            for row in reader:
                self.P1_ativo.append(row)
        
        
        P2_df = Workbook()
        self.P2_ativo = P2_df.active
        with open('P2 Outubro 2023.csv') as f:
            reader = csv.reader(f,delimiter=";")
            for row in reader:
                self.P2_ativo.append(row)

        
        
        return self.screen
    

    def P1(self):
        self.tipo = 1
        print('p1')
    def P2(self):
        self.tipo = 2
        print('p2')
        
    
    def mostraInfo(self):
        linha= 0
        contrato = (self.root.get_screen('tela').ids.contrato.text )
        if self.tipo == 1:
            excel = self.P1_ativo
        else:
            excel = self.P2_ativo
        
        for row in excel:
            
            linha+=1
            if row[0].value != 'DATA_ATIVACAO' and len(row[19].value) != 0:

                valor = str(int (float(row[19].value.replace(',','.'))))
                print(valor)
                if valor == contrato:
                    self.root.get_screen('tela').ids.linha.text = str(linha)
                    self.root.get_screen('tela').ids.data.text = str (row[0].value[0:10])
                    self.root.get_screen('tela').ids.franquia.text = str(row[3].value)
                    self.root.get_screen('tela').ids.comicao.text = str(row[7].value)
                    print(row[2].value,'/valor franquia',row[12].value)
                    


    
            
            
            
if __name__ == '__main__':
    procuraContrato().run()