import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv

dados_cidades = {
    "Alegrete": '390',
    "Almirante Tamandaré": '294',
    "Alvorada": '458',
    "Apucarana": '322',
    "Arapongas": '022',
    "Araranguá": '400',
    "Araucária": '333',
    "Arroio do Meio": '618',
    "Bagé": '690',
    "Balneário Camboriú": '752',
    "Bento Gonçalves": '688',
    "Biguaçu": '058',
    "Blumenau": '700',
    "Brusque": '070',
    "Caçador": '819',
    "Cachoeira do Sul": '394',
    "Cachoeirinha": '560',
    "Camaquã": '466',
    "Cambé": '379',
    "Camboriú": '868',
    "Campo Bom": '584',
    "Campo Largo": '389',
    "Canela": '601',
    "Canoas": '603',
    "Capão da Canoa": '87',
    "Carazinho": '614',
    "Carlos Barbosa": '920',
    "Cascavel": '641',
    "Caxias do Sul": '687',
    "Chapecó": '068',
    "Charqueadas": '217',
    "Cianorte": '430',
    "Colombo": '437',
    "Concórdia": '113',
    "Criciúma": '89',
    "Cruz Alta": '685',
    "Curitiba": '884',
    "Dois Irmãos": '781',
    "Eldorado do Sul": '918',
    "Encantado": '934',
    "Erechim": '695',
    "Estância Velha": '757',
    "Esteio": '758',
    "Estrela": '231',
    "Farroupilha": '83',
    "Florianópolis": '88',
    "Foz do Iguaçu": '506',
    "Fraiburgo": '449',
    "Frederico Westphalen": '536',
    "Garibaldi": '568',
    "Gaspar": '154',
    "Gramado": '817',
    "Gravataí": '720',
    "Guaíba": '728',
    "Guaramirim": '163',
    "Guarapuava": '531',
    "Herval d'Oeste": '698',
    "Ibiporã": '204',
    "Içara": '724',
    "Igrejinha": '885',
    "Imbé": '835',
    "Indaial": '775',
    "Itajaí": '193',
    "Itapema": '94',
    "Itaqui": '106',
    "Ivoti": '173',
    "Jaraguá do Sul": '203',
    "Joaçaba": '119',
    "Joinville": '086',
    "Lages": '213',
    "Lajeado": '689',
    "Londrina": '996',
    "Mafra": '348',
    "Marau": '638',
    "Maringá": '091',
    "Montenegro": '950',
    "Navegantes": '260',
    "Nova Petrópolis": '201',
    "Novo Hamburgo": '686',
    "Osório": '293',
    "Palhoça": '278',
    "Palmas": '924',
    "Palmeira das Missões": '412',
    "Panambi": '447',
    "Paranaguá": '749',
    "Parobé": '499',
    "Passo Fundo": '693',
    "Pelotas": '691',
    "Pinhais": '770',
    "Piraquara": '380',
    "Ponta Grossa": '794',
    "Porto Alegre": '78',
    "Rio do Sul": '301',
    "Rio Grande": '692',
    "Rio Negrinho": '352',
    "Rio Pardo": '478',
    "Rolândia": '851',
    "Rosário do Sul": '613',
    "Sant'Ana do Livramento": '041',
    "Santa Cruz do Sul": '694',
    "Santa Maria": '075',
    "Santa Rosa": '940',
    "Santo Ângelo": '77',
    "São Bento do Sul": '570',
    "São Borja": '229',
    "São Francisco do Sul": '620',
    "São Gabriel": '270',
    "São José": '363',
    "São José dos Pinhais": '923',
    "São Leopoldo": '710',
    "São Lourenço do Sul": '482',
    "São Luiz Gonzaga": '504',
    "Sapiranga": '330',
    "Sapucaia do Sul": '332',
    "Taquara": '242',
    "Telêmaco Borba": '331',
    "Teutônia": '324',
    "Timbó": '026',
    "Torres": '398',
    "Tramandaí": '438',
    "Três Coroas": '497',
    "Tubarão": '090',
    "União da Vitória": '593',
    "Uruguaiana": '684',
    "Vacaria": '764',
    "Venâncio Aires": '829',
    "Vera Cruz": '839',
    "Veranópolis": '853',
    "Viamão": '871',
    "Videira": '236',
    "Xanxerê": '334',
    "Xaxim": '361',
}

P2 = Workbook()
P2_ativo = P2.active

P1 = Workbook()
P1_ativo = P1.active

with open('P2 Novembro 2023.csv') as f:
    reader = csv.reader(f,delimiter=";")

    for row in reader:
        P2_ativo.append(row)
        
        
with open('P1 Novembro 2023.csv') as f:
    reader = csv.reader(f,delimiter=";")

    for row in reader:
        P1_ativo.append(row)



        
           
        
        
df_Grupo = load_workbook('grupo_e_gian.xlsx')



have_name = False
have_contrato = False

wbp1 = Workbook()
wap1 = wbp1.worksheets[0]
wap1['A1']='Vendedor'
wap1['B1']='Total comição'
wap1['C1']='Total franquia'
wap1['D1']='Número de contratos'
wap1['E1']='Número de estornos'

wap1['B2']='0'
wap1['C2']='0'
wap1['D2']='0'
wap1['E2']='0'


wbp2 = Workbook()
wap2 = wbp2.worksheets[0]
wap2['A1']='Vendedor'
wap2['B1']='Total comição'
wap2['C1']='Total franquia'
wap2['D1']='Número de contratos'
wap2['E1']='Número de estornos'

wap2['B2']='0'
wap2['C2']='0'
wap2['D2']='0'
wap2['E2']='0'

wb_contrato = Workbook()
wb_c_ativo = wb_contrato.active



#row contrato = 12
#row vendedor = 14
#row cidades = 16
#row comição = 7
#


contador = 0
df_Grupoa = df_Grupo.active

for row in df_Grupoa:
    have_contrato = False

    have_name = False
    
    if row[0].value == None:
        pass
    
    else:
        contrato_completo = row[1].value
        #print(row[14].value)
        #print(contrato_completo)

        for rows in P1_ativo:

            #####P1
            
            if rows[19].value == contrato_completo and rows[0].value[0:10] == row[2].value:
                
                have_contrato = True
                #print('valor franquia',rows[12].value,contrato_completo, rows[0].value)
                
                for nome in wap1:
                    
                    if row[0].value == nome[0].value:
                        have_name = True

                        break
                
                
                if have_name:
                    
                    for linha in wap1:

                        if row[0].value == linha[0].value:
                            
                            valor = float (linha[1].value)
                            
                            valor +=float (rows[7].value.replace(',','.'))
                            linha[1].value = valor
                            
                            valor = float (linha[2].value)
                            valor += float (rows[12].value.replace(',','.'))
                            linha[2].value = valor
                            linha[3].value += 1
                            
                            break
                else:
                    for linha in wap1:
                        
                        if linha[0].value == None:
                            linha[0].value = row[0].value
                            linha[1].value = float(rows[7].value.replace(',','.'))
                            linha[2].value = float(rows[12].value.replace(',','.'))
                            linha[3].value = 1
                            wap1.insert_rows(2)
                            
                            break
                break
            ####P2
        have_name = False
        for rows in P2_ativo:
            
            print(rows[0].value[0:10] ,'             ' ,row[2].value)
            if rows[19].value == contrato_completo and rows[0].value[0:10] == row[2].value:
                
                
                have_contrato = True
                #print('valor franquia',rows[12].value,contrato_completo, rows[0].value)
                
                for nome in wap2:
                    
                    if row[0].value == nome[0].value:
                        have_name = True

                        break
                
                
                if have_name:
                    
                    for linha in wap2:

                        if row[0].value == linha[0].value:
                            
                            valor = float (linha[1].value)
                            
                            valor +=float (rows[7].value.replace(',','.'))
                            linha[1].value = valor
                            
                            valor = float (linha[2].value)
                            valor += float (rows[12].value.replace(',','.'))
                            linha[2].value = valor
                            linha[3].value += 1
                          
                            break
                else:
                    for linha in wap2:
                        
                        if linha[0].value == None:
                            linha[0].value = row[0].value
                            linha[1].value = float(rows[7].value.replace(',','.'))
                            linha[2].value = float(rows[12].value.replace(',','.'))
                            linha[3].value = 1
                            wap2.insert_rows(2)
                            
                            
                            
                            break
                break
                
        if have_contrato:
            
            contador+=1
        
        else:
            ultimalinha = len(wb_c_ativo['A'])+1
            wb_c_ativo[f'A{ultimalinha}']=row[0].value
            wb_c_ativo[f'B{ultimalinha}']=contrato_completo
            wb_c_ativo[f'C{ultimalinha}']=row[2].value
            
                    
                    
            
        
print(contador)
df_estornos = Workbook()



estorno = df_estornos.active

estorno['A1']='Vendedor'
estorno['B1']='Contrato'
estorno['C1']='Data'
estorno['D1']='Comição'
estorno['E1']='Comição com fator'
estorno['F1']='Fator'
estorno['G1']='P'



df_banco = load_workbook('banco_contratos.xlsx')
banco = df_banco.active

for linha_P in P1_ativo:
    if linha_P[2].value[0:7] == 'ESTORNO':
        
        #print(linha_P[19].value)
        for linha_banco in banco:

            if linha_banco[1].value == linha_P[19].value and  linha_P[0].value[0:10]== linha_banco[2].value :
                a = len(estorno['A'])+1

                vendedor = linha_banco[0].value
                contrato = linha_banco[1].value
                data = linha_P[0].value[0:10]     
                comicao = float (linha_P[7].value.replace(',','.'))
                fator = float(linha_P[23].value.replace(',','.'))
                comicao_com_fator = comicao / fator
                
                
                estorno[f'A{a}']=vendedor
                estorno[f'B{a}']=contrato
                estorno[f'C{a}']=data
                estorno[f'D{a}']=comicao
                estorno[f'E{a}']=comicao_com_fator
                estorno[f'F{a}']=fator

                estorno[f'G{a}']='P1'

                           



for linha_P in P2_ativo:
    if linha_P[2].value[0:7] == 'ESTORNO':
        for linha_banco in banco:
            
            if linha_banco[1].value == linha_P[19].value and  linha_P[0].value[0:10]== linha_banco[2].value :
                
                a = len(estorno['A'])+1

                vendedor = linha_banco[0].value
                contrato = linha_banco[1].value
                data = linha_P[0].value[0:10]     
                comicao = float (linha_P[7].value.replace(',','.'))
                fator = float(linha_P[23].value.replace(',','.'))
                comicao_com_fator = comicao / fator
                
                
                estorno[f'A{a}']=vendedor
                estorno[f'B{a}']=contrato
                estorno[f'C{a}']=data
                estorno[f'D{a}']=comicao
                estorno[f'E{a}']=comicao_com_fator
                estorno[f'F{a}']=fator
                estorno[f'G{a}']='P2'










df_estornos.save('estornos.xlsx')
wbp1.save('dados p1.xlsx')
wbp2.save('dados p2.xlsx')
wb_contrato.save('contratos não encontrados.xlsx')





    




