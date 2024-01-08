import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from unidecode import unidecode

from kivy.lang import Builder

from kivymd.app import MDApp
import os
from kivy.clock import Clock
import shutil
import threading
from kivymd.uix.screen import MDScreen
from kivymd.app import MDApp
from kivy.uix.image import Image
from kivymd.uix.button import MDFillRoundFlatIconButton, MDFillRoundFlatButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivymd.uix.toolbar import MDToolbar
from kivy.config import Config
from kivy.app import App
from kivy import platform
from kivymd.uix.filemanager import MDFileManager 
from kivymd.uix.selectioncontrol import MDCheckbox
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from datetime import date
from tkinter.filedialog import askopenfilename
from kivy.config import Config
from kivymd.toast.kivytoast.kivytoast import toast

Config.set('input', 'mouse', 'mouse,disable_multitouch')
screen = '''

ScreenManager:
    Tela:
    
<Tela>:
    name:'tela'
    
    MDToolbar:
        title: "banco de dados"
        pos_hint: {"top" : 1}
        

    MDTextField:
        id: grupo
        size_hint_x:(0.35)
        font_size : 15
        pos_hint : {"center_x":0.25,"center_y":0.7}
        hint_text: "grupo"
    MDIconButton:
        icon: "arrow-up"
        pos_hint: {"center_x": .45, "center_y": .7}
        on_press : app.nome_arquivos('grupo')
        
    MDTextField:
        id: gian
        size_hint_x:(0.35)
        font_size : 15
        pos_hint : {"center_x":0.70,"center_y":0.7}
        hint_text: "gian"
    MDIconButton:
        icon: "arrow-up"
        pos_hint: {"center_x": .9, "center_y": .7}
        on_press : app.nome_arquivos('gian')
        
    MDTextField:
        id: P1
        size_hint_x:(0.35)
        font_size : 15
        pos_hint : {"center_x":0.25,"center_y":0.5}
        hint_text: "P1"
    MDIconButton:
        icon: "arrow-up"
        pos_hint: {"center_x": .45, "center_y": .5}
        on_press : app.nome_arquivos('P1')
        
    MDTextField:
        id: P2
        size_hint_x:(0.35)
        font_size : 15
        pos_hint : {"center_x":0.70,"center_y":0.5}
        hint_text: "P2"
    MDIconButton:
        icon: "arrow-up"
        pos_hint: {"center_x": .9, "center_y": .5}
        on_press : app.nome_arquivos('P2')
    
    MDFillRoundFlatButton:
        text:'Adicionar ao banco de dados'
        pos_hint: {"center_x": .7, "center_y": .2}
        on_press : app.rodadados_em_thread_banco() 
        size_hint_y:(0.2)
        
        
    MDFillRoundFlatButton:
        text:'Adicionar ao banco de dados'
        pos_hint: {"center_x": .3, "center_y": .2}
        on_press : app.rodadados_em_thread_organiza() 
        size_hint_y:(0.2)
        

        
'''

class Tela(Screen):
    pass


sm = ScreenManager()
sm.add_widget(Tela(name='tela'))

class claroApp(MDApp):
    def build(self):
        self.screen=Builder.load_string(screen)


        return self.screen
    def nome_arquivos(self,id):
        
        file = askopenfilename()
        toast('arquivo selecionado')
        if id == 'gian':
            self.root.get_screen('tela').ids.gian.text = file
        elif id == 'grupo':
            self.root.get_screen('tela').ids.grupo.text = file
        elif id == 'P1':
            self.root.get_screen('tela').ids.P1.text = file
        elif id == 'P2':
            self.root.get_screen('tela').ids.P2.text = file
    
    
    
    
    def rodadados_em_thread_banco(self,value):
        print('a')
        self.tipo = 'banco'
    

        
        t=threading.Thread(target=self.criaDados)
        t.start()

    def rodadados_em_thread_organiza(self):
        print('a')
        self.tipo = 'organizador'
    

        
        t=threading.Thread(target=self.organiza_dados)
        t.start()


    def criaDados(self):  
    
        dados_cidades = {
        "ADAMANTINA": '323',
        "ALVORADA SUL": '021',
        "ALVORADO SUL": '021',
        'ALVORADA DO SUL':'021',
        "AGUDOS": '411',
        "ALAGOINHAS": '802',
        "ALEGRETE": '390',
        "ALMIRANTE TAMANDARE": '294',
        "ALUMINIO": '535',
        "ALVARES MACHADO": '551',
        "ALVORADA": '458',
        "AMERICANA": '386',
        "AMERICO BRASILIENSE": '624',
        "AMPARO": '538',
        "ANANINDEUA": '178',
        "ANAPOLIS": '015',
        "ANDRADINA": '697',
        "APARECIDA": '399',
        "APARECIDA DE GOIANIA": '895',
        "APUCARANA": '322',
        "AQUIRAZ": '561',
        "ARACAJU": '550',
        "ARACATUBA": '404',
        "ARACRUZ": '157',
        "ARAGUAINA": '679',
        "ARAGUARI": '455',
        "ARAPIRACA": '214',
        "ARAPONGAS": '093',
        "ARARANGUA": '400',
        "ARARAQUARA": '409',
        "ARARAS": '413',
        "ARAUCARIA": '333',
        "ARAXA": '151',
        "ARIQUEMES": '870',
        "ARMACAO DOS BUZIOS": '116',
        "ARROIO DO MEIO": '618',
        "ARROIO DO MEIA": '618',
        "ARTUR NOGUEIRA": '419',
        "ARUJA": '420',
        "ATIBAIA": '423',
        "AVARE": '092',
        "BADY BASSITT": '124',
        "BAGE": '690',
        "BALNEARIO CAMBORIU": '752',
        "BARBACENA": '470',
        "BARRA MANSA": '100',
        "BARREIRAS": '353',
        "BARRETOS": '302',
        "BARRINHA": '442',
        "BARUERI": '443',
        "BATATAIS": '448',
        "BAURU": '008',
        "BEBEDOURO": '388',
        "BELEM": '194',
        "BELFORD ROXO": '105',
        "BELO HORIZONTE": '013',
        "BENTO GONCALVES": '688',
        "BERTIOGA": '456',
        "BETIM": '514',
        "BIGUACU": '058',
        "BIRIGUI": '450',
        "BLUMENAU": '700',
        "BOITUVA": '464',
        "BOTUCATU": '471',
        "BRAGANCA PAULISTA": '474',
        "BRASILIA": '040',
        "BRUSQUE": '070',
        "CABEDELO": '133',
        "CABO FRIO": '115',
        "CABREUVA": '787',
        "CACADOR": '819',
        "CACAPAVA": '487',
        "CACHOEIRA DO SUL": '394',
        "CACHOEIRA PAULISTA": '488',
        "CACHOEIRINHA": '560',
        "CACHOEIRO DE ITAPEMIRIM": '911',
        "CACOAL": '898',
        "CAIEIRAS": '498',
        "CAJAMAR": '886',
        "CALDAS NOVAS": '073',
        "CAMACARI": '750',
        "CAMAQUA": '466',
        "CAMBE": '379',
        "CAMBORIU": '868',
        "CAMBURIU": '868',
        "CAMPINA GRANDE": '847',
        "CAMPINAS": '052',
        "CAMPO BOM": '584',
        "CAMPO GRANDE": '011',
        "CAMPO LARGO": '389',
        "CAMPO LIMPO PAULISTA": '012',
        "CAMPOS DO JORDAO": '028',
        "CAMPOS DOS GOYTACAZES": '529',
        "CANELA": '601',
        "CANOAS": '603',
        "CAPAO DA CANOA": '087',
        "CAPAO DA CANOS":'087',
        "CAPIVARI": '530',
        "CARAGUATATUBA": '553',
        "CARAPICUIBA": '533',
        "CARAZINHO": '614',
        "CARIACICA": '917',
        "CARLOS BARBOSA": '920',
        "CARUARU": '152',
        "CASA BRANCA": '202',
        "CASCAVEL": '641',
        "CASTANHAL": '023',
        "CATAGUASES": '340',
        "CATANDUVA": '543',
        "CAXIAS DO SUL": '687',
        "CERQUILHO": '296',
        "CHAPECO": '068',
        "CHARQUEADAS": '217',
        "CIANORTE": '430',
        "COLATINA": '489',
        "COLOMBO": '437',
        "CONCORDIA": '113',
        "CONSELHEIRO LAFAIETE": '741',
        "CONTAGEM": '753',
        "CORDEIROPOLIS": '432',
        "CORONEL FABRICIANO": '778',
        "COSMOPOLIS": '565',
        "COTIA": '567',
        "CRAVINHOS": '569',
        "CRICIUMA": '089',
        "CRUZ ALTA": '685',
        "CRUZEIRO": '572',
        "CUBATAO": '573',
        "CUIABA": '719',
        "CURITIBA": '884',
        "DESCALVADO": '576',
        "DIADEMA": '577',
        "DIVINOPOLIS": '832',
        "DOIS IRMAOS": '781',
        "DOURADOS": '867',
        "DRACENA": '586',
        "DUQUE DE CAXIAS": '161',
        "ELDORADO DO SUL": '918',
        "ELIAS FAUSTO": '595',
        "EMBU DAS ARTES": '599',
        "ENCANTADO": '934',
        "ERECHIM": '695',
        "ESPIRITO SANTO DO PINHAL": '970',
        "ESTANCIA VELHA": '757',
        "ESTEIO": '758',
        "ESTRELA": '231',
        "EUNAPOLIS": '061',
        "EUSEBIO": '044',
        "FARROUPILHA": '083',
        "FEIRA DE SANTANA": '852',
        "FERNANDOPOLIS": '098',
        "FLORIANOPOLIS": '088',
        "FLORIPA":'088',
        "FORMOSA": '782',
        "FORTALEZA": '097',
        "FOZ DO IGUACU": '506',
        "FRAIBURGO": '449',
        "FRANCA": '056',
        "FREDERICO WESTPHALEN": '536',
        "GARCA": '232',
        "GARIBALDI": '568',
        "GASPAR": '154',
        "GOIANIA": '010',
        "GOVERNADOR VALADARES": '930',
        "GRAMADO": '817',
        "GRAVATAI": '720',
        "GUAIBA": '728',
        "GUAIRA": '364',
        "GUAPIACU": '381',
        "GUARAMIRIM": '163',
        "GUARAPUAVA": '531',
        "GUARARAPES": '496',
        "GUARATINGUETA": '656',
        "GUARATUBA": '081',
        "GUARUJA": '659',
        "GUARULHOS": '661',
        "GURUPI": '992',
        "HERVAL DO OESTE": '698',
        "HERVAL D OESTE": '698',
        "HORTOLANDIA": '668',
        "IBATE": '674',
        "IBIPORA": '204',
        "IBIUNA": '755',
        "ICARA": '724',
        "IGREJINHA": '885',
        "ILHEUS": '843',
        "IMBE": '835',
        "IMPERATRIZ": '051',
        "INDAIAL": '775',
        "INDAIATUBA": '054',
        "IPATINGA": '405',
        "IPERO": '951',
        "ITABIRA": '813',
        "ITABUNA": '084',
        "ITAGUAI": '057',
        "ITAJAI": '193',
        "ITAJUBA": '402',
        "ITANHAEM": '712',
        "ITAPECERICA DA SERRA": '714',
        "ITAPEMA": '094',
        "ITAPETININGA": '715',
        "ITAPEVA": '166',
        "ITAPEVI": '722',
        "ITAPEIRA": '183',
        "ITAQUAQUECETUBA": '733',
        "ITAQUI": '106',
        "ITATIBA": '739',
        "ITAUNA": '144',
        "ITU": '746',
        "ITUIUTABA": '060',
        "ITUMBIARA": '257',
        "ITUPEVA": '360',
        "ITUVERAVA": '574',
        "IVOTI": '173',
        "JABOATAO DOS GUARARAPES": '223',
        "JABOTICABAL": '410',
        "JACAREI": '756',
        "JAGUARIUNA": '761',
        "JALES": '492',
        "JANDIRA": '765',
        "JARAGUA DO SUL": '203',
        "JARDINOPOLIS": '575',
        "JARINU": '581',
        "JATAI": '303',
        "JAU": '769',
        "JEQUIE": '718',
        "JI-PARANA": '994',
        "JOACABA": '119',
        "JOAO MONLEVADE": '526',
        "JOAO PESSOA": '907',
        "JOINVILLE": '086',
        "JOIVILLE": '086',
        "JOSE BONIFACIO": '662',
        "JUAZEIRO": '823',
        "JUAZEIRO DO NORTE": '110',
        "JUIZ DE FORA": '126',
        "JUNDIAI": '055',
        "LAGES": '213',
        "LAGE": '213',
        "LAGOA SANTA": '043',
        "LAJEADO": '689',
        "LARANJAL PAULISTA": '859',
        "LAURO DE FREITAS": '042',
        "LAVRAS": '909',
        "LEME": '905',
        "LENCOIS PAULISTA": '913',
        "LIMEIRA": '791',
        "LINS": '948',
        "LONDRINA": '996',
        "LORENA": '795',
        "LOUVEIRA": '798',
        "LUCAS DO RIO VERDE": '875',
        "MACAE": '197',
        "MACAPA": '393',
        "MACEIO": '480',
        "MAFRA": '348',
        "MAIRINQUE": '810',
        "MANAUS": '121',
        "MANHUACU": '222',
        "MARABA": '002',
        "MARAU": '638',
        "MARECHAL CANDIDO RONDON": '208',
        "MARILIA": '820',
        "MARINGA": '091',
        "MATA DE SAO JOAO": '481',
        "MATAO": '297',
        "MAUA": '833',
        "MEDIANEIRA": '428',
        "MESQUITA": '239',
        "MIGUEL PEREIRA": '391',
        "MIRANDOPOLIS": '392',
        "MIRASSOL": '850',
        "MOCOCA": '431',
        "MOGI DAS CRUZES": '856',
        "MOGI GUACU": '863',
        "MOGI MIRIM": '865',
        "MONGAGUA": '869',
        "MONTE ALTO": '534',
        "MONTE MOR": '882',
        "MONTENEGRO": '950',
        "MONTES CLAROS": '275',
        "MORUNGABA": '626',
        "MOSSORO": '947',
        "NATAL": '095',
        "NAVEGANTES": '260',
        "NILOPOLIS": '226',
        "NITEROI": '228',
        "NOVA FRIBURGO": '030',
        "NOVA IGUACU": '237',
        "NOVA LIMA": '312',
        "NOVA MUTUM": '341',
        "NOVA ODESSA": '910',
        "NOVA PETROPOLIS": '201',
        "NOVO HAMBURGO": '686',
        "OLIMPIA": '987',
        "OLINDA": '269',
        "ORLANDIA": '592',
        "OSASCO": '924',
        "OSORIO": '293',
        "OURINHOS": '071',
        "PAICANDU": '912',
        'PALHOCA': '278',
        'PALMAS': '540',
        'PALMEIRA DAS MISSOES': '412',
        'PANAMBI': '447',
        'PARA DE MINAS': '812',
        'PARAGOMINAS': '628',
        'PARAIBA DO SUL': '677',
        'PARAISO DO TOCANTINS': '085',
        'PARANAGUA': '749',
        'PARAUAPEBAS': '745',
        'PARNAIBA': '069',
        'PARNAMIRIM': '637',
        'PAROBE': '499',
        'PASSO FUNDO': '693',
        'PASSOS': '995',
        'PATOS DE MINAS': '014',
        'PAULINIA': '954',
        'PAULISTA': '291',
        'PEDREIRA': '966',
        'PEDRO LEOPOLDO': '234',
        'PELOTAS': '691',
        'PENAPOLIS': '495',
        'PERUIBE': '972',
        'PETROLINA': '956',
        'PETROPOLIS': '250',
        'PIEDADE': '571',
        'PINDAMONHANGABA': '976',
        'PINHAIS': '770',
        'PIRACAIA': '657',
        'PIRACICABA': '009',
        'PIRAQUARA': '380',
        'PIRASSUNUNGA': '747',
        'POA': '609',
        'POCOS DE CALDAS': '037',
        'PONTA GROSSA': '794',
        'PONTA PORA': '017',
        'PONTAL': '844',
        'PORTO ALEGRE': '078',
        'PORTO FELIZ': '022',
        'PORTO FERREIRA': '908',
        'PORTO SEGURO': '532',
        'PORTO VELHO': '220',
        'POTIM': '024',
        'POTIRENDABA': '941',
        'POUSO ALEGRE': '459',
        'PRAIA GRANDE': '027',
        'PRESIDENTE BERNARDES': '025',
        'PRESIDENTE PRUDENTE': '036',
        'PROMISSAO': '074',
        'QUATRO BARRAS': '683',
        'RAFARD': '048',
        'RECIFE': '136',
        'REGISTRO': '211',
        'RESENDE': '268',
        'RIBEIRAO PIRES': '067',
        'RIBEIRAO PRETO': '005',
        'RIO BRANCO': '066',
        'RIO CLARO': '076',
        'RIO DAS OSTRAS': '285',
        'RIO DE JANEIRO': '038',
        'RIO DO SUL': '301',
        'RIO GRANDE': '692',
        'RIO NEGRINHO': '352',
        'RIO PARDO': '478',
        'RIO VERDE': '112',
        'ROLANDIA': '851',
        'RONDONOPOLIS': '818',
        'ROSARIO DO SUL': '613',
        'SABARA': '541',
        'SALTO': '295',
        'SALVADOR': '230',
        'SANTA BARBARA DO OESTE': '103',
        'SANTA CRUZ DO RIO PARDO': '108',
        'SANTA CRUZ DO SUL': '694',
        'SANTA CRUZ': '694',
        'SANTA GERTRUDES': '793',
        'SANTA HELENA DE GOIAS': '580',
        'SANTA ISABEL': '808',
        'SANTA LUZIA': '072',
        'SANTA MARIA': '075',
        'SANTA ROSA': '940',
        'SANTA ROSA DE VITERBO': '904',
        'SANTANA': '616',
        'SANTANA DE PARNAIBA': '621',
        'SANT ANA DO LIVRAMENTO': '041',
        'SANTO ANDRE': '129',
        'SANTO ANGELO': '077',
        'SANTOS': '004',
        'SAO BENTO': '570',
        'SAO BENTO ': '570',
        'SAO BERNARDO DO CAMPO': '141',
        'SAO BORJA': '229',
        'SAO CAETANO DO SUL': '143',
        'SAO CARLOS': '053',
        'SAO FRANCISCO DO SUL': '620',
        'SAO GABRIEL': '270',
        'SAO GONCALO': '305',
        'SAO JOAO DA BOA VISTA': '200',
        'SAO JOAO DE MERITI': '316',
        'SAO JOAO DEL REI': '273',
        'SAO JOAQUIM DA BARRA': '615',
        'SAO JOSE': '363',
        'SAO JOSE DO RIO PARDO': '315',
        'SAO JOSE DO RIO PRETO': '006',
        'SAO JOSE DOS CAMPOS': '162',
        'SAO JOSE DOS PINHAIS': '923',
        'SAO LEOPOLDO': '710',
        'SAO LOURENCO DO SUL': '482',
        'SAO LUIS': '096',
        'SAO LUIZ GONZAGA': '504',
        'SAO PAULO': '003',
        'SAO PEDRO DA ALDEIA': '320',
        'SAO ROQUE': '678',
        'SAO SEBASTIAO': '182',
        'SAO VICENTE': '187',
        'SAPIRANGA': '330',
        'SAPUCAIA DO SUL': '332',
        'SENADOR CANEDO': '866',
        'SERRA': '566',
        'SERRA NEGRA': '579',
        'SERRANA': '587',
        'SERTAOZINHO': '195',
        'SETE LAGOAS': '742',
        'SINOP': '837',
        'SOBRAL': '521',
        'SOROCABA': '007',
        'SORRISO': '838',
        'SUMARE': '207',
        'SUZANO': '209',
        'TABOAO DA SERRA': '216',
        'TAMBAU': '854',
        'TAQUARA': '242',
        'TATUI': '233',
        'TAUBATE': '235',
        'TEIXEIRA DE FREITAS': '807',
        'TELEMACO BORBA': '331',
        'TELEMACO': '331',
        'TEOFILO OTONI': '763',
        'TERESINA': '063',
        'TERESOPOLIS': '338',
        'TEUTONIA': '324',
        'TIETE': '241',
        'TIMBO': '026',
        'TIMON': '165',
        'TIMOTEO': '477',
        'TORRES': '398',
        'TRAMANDAI': '438',
        'TRAMADAI': '438',
        'TREMEMBE': '246',
        'TRES CORACOES': '780',
        'TRES COROAS': '497',
        'TRES LAGOAS': '378',
        'TRES RIOS': '771',
        'TRINDADE': '047',
        'TUBARAO': '090',
        'UBA': '164',
        'UBATUBA': '261',
        'UBERABA': '803',
        'UBERLANDIA': '806',
        'UNIAO DA VITORIA': '593',
        'UNIAO DA V': '593',
        'UNIAO': '593',
        'UNIAO ': '593',
        'URUGUAIANA': '684',
        'VACARIA': '764',
        'VALENCA': '800',
        'VALINHOS': '272',
        'VALPARAISO': '362',
        'VALPARAISO DE GOIAS': '160',
        'VARGEM GRANDE PAULISTA': '276',
        'VARGINHA': '821',
        'VARZEA GRANDE': '858',
        'VARZEA PAULISTA': '277',
        'VASSOURAS': '836',
        'VENANCIO AIRES': '829',
        'VENANCIO': '829',
        'VENANCIO ': '829',
        'VERA CRUZ': '839',
        'VERANOPOLIS': '853',
        'VESPASIANO': '834',
        'VIAMAO': '871',
        'VICOSA': '588',
        'VIDEIRA': '236',
        'VILA VELHA': '507',
        'VILHENA': '206',
        'VINHEDO': '279',
        'VITORIA': '508',
        'VITORIA DA CONQUISTA': '336',
        'VOLTA REDONDA': '359',
        'VOTORANTIM': '282',
        'VOTUPORANGA': '559',
        'XANXERE': '334',
        'XAXIM': '361',
        'JARAGUA': '203',
        'SAO BENTO DO SUL':'000',
        'PORTO UNIAO':'000'
    }

        self.gian = self.root.get_screen('tela').ids.gian.text
        self.grupo = self.root.get_screen('tela').ids.grupo.text
        

        df_Grupo = load_workbook(self.grupo, keep_vba=True, data_only=True)
        gian = load_workbook(self.gian, keep_vba=True, data_only=True)
        ag=gian.worksheets[0]
        
        if self.tipo == 'banco':
            fusao = load_workbook('banco_contratos.xlsx')
            
        elif self.tipo == 'organizador':
            fusao = Workbook()
            
        f_ativo = fusao.active

        tabela = 0 
        for grupo in df_Grupo.worksheets:

            
            for i in range(1,26):
                df = df_Grupo.worksheets[tabela]
                
                if df.cell(row=1,column=i).value == 'Dia inst':
                    data_instalacao_row=i-1
                elif df.cell(row=1,column=i).value == 'Contrato':
                    contrato_row=i-1
                elif df.cell(row=1,column=i).value == 'Instalado?':
                    instalado_row=i-1
                elif df.cell(row=1,column=i).value == 'Vendedor':
                    vendedor_row=i-1
                elif df.cell(row=1,column=i).value == 'Cidade':
                    cidade_row=i-1
                elif df.cell(row=1,column=i).value == 'Cod Cidade':
                    cod_cidade_row=i-1


            for row in grupo:

                
                
                if row[0].value == None or grupo.title == 'dados':
                        break
                
                elif row[cidade_row].value == 'Cidade' or row[contrato_row].value == '---' or row[contrato_row].value == None or row[instalado_row].value != 'INSTALADA' or row[data_instalacao_row].value == None or (row[cidade_row].value == None and (row[cod_cidade_row].value == None or row[cod_cidade_row].value == '#N/A')) :
                    pass    
                
                else:
                    
                    if row[cod_cidade_row].value != None and row[cod_cidade_row].value != '#N/A':
                        cod = str (int (row[cod_cidade_row].value))
                        
                        while  len(cod) <3:
                            cod = '0'+cod
                        contrato_completo = str(cod)+str(row[contrato_row].value)
                        print(contrato_completo)
                    else:
                        
                        contrato_completo = str (dados_cidades [unidecode(row[cidade_row].value.upper())])+str(row[contrato_row].value)
                        print(dados_cidades [unidecode(row[cidade_row].value.upper())])
                        print(row[cidade_row].value)
                        print(contrato_completo)

                    vendedor = row[vendedor_row].value
                    data = str(row[data_instalacao_row].value)[0:10].replace('/','-')
                    ano,mes,dia = data.split('-')
                    print(len(f_ativo['A']))
                    ultimalinha = len(f_ativo['A'])+1
                    f_ativo[f'A{ultimalinha}']=vendedor
                    f_ativo[f'B{ultimalinha}']=contrato_completo
                    
                    if len(ano) == 4:
                        
                        f_ativo[f'C{ultimalinha}']=f'{ano}-{mes}-{dia}'
                        
                    else:
                        
                        f_ativo[f'C{ultimalinha}']=f'{dia}-{mes}-{ano}' 
                        
            tabela+= 1
                        
                        
        for row in ag:
            
            for i in range(1,26):
                
                if ag.cell(row=1,column=i).value == 'DATA DA INSTALAÇÃO' or ag.cell(row=1,column=i).value == 'Dia inst':
                    data_instalacao_row=i-1
                elif ag.cell(row=1,column=i).value == 'CONTRATO' or ag.cell(row=1,column=i).value == 'Contrato':
                    contrato_row=i-1
                elif ag.cell(row=1,column=i).value == 'INSTALADO' or ag.cell(row=1,column=i).value == 'Instalado?':
                    instalado_row=i-1
                elif ag.cell(row=1,column=i).value == 'CIDADE' or ag.cell(row=1,column=i).value == 'Cidade':
                    cidade_row=i-1

            
            if row[cidade_row].value == 'Cidade' or row[cidade_row].value == '---' or row[cidade_row].value == None or row[cidade_row].value == None or (row[instalado_row].value != 'ok' and row[instalado_row].value != 'INSTALADA') or row[data_instalacao_row].value == None or (row[cidade_row].value == None and (row[cod_cidade_row].value == None or row[cod_cidade_row].value == '#N/A')):

                pass
            elif row[0].value == None:
                break
                
            else:
                
                contrato_completo = str (dados_cidades [unidecode(row[cidade_row].value.upper())])+str(row[contrato_row].value)
                data = str(row[data_instalacao_row].value)[0:10].replace('/','-')
                dia,mes,ano = data.split('-')
                ultimalinha = len(f_ativo['A'])+1
                f_ativo[f'A{ultimalinha}']='GIAN'
                f_ativo[f'B{ultimalinha}']=contrato_completo
                if len(ano) == 4:
                        
                    f_ativo[f'C{ultimalinha}']=f'{ano}-{mes}-{    dia}'
                        
                else:
                        
                    f_ativo[f'C{ultimalinha}']=f'{dia}-{mes}-{ano}' 
                
                
        #data instalação = 10
        #contrato = 13
        #cidade = 17        
        
        if self.tipo == 'banco':
            fusao.save('banco_contratos.xlsx')
            toast('Adicionado')

        elif self.tipo == 'organizador':
            return fusao
                        
                        
                        
        
                    
      
    def organiza_dados(self):
        
        P2 = Workbook()
        P2_ativo = P2.active

        P1 = Workbook()
        P1_ativo = P1.active

        self.P1 = self.root.get_screen('tela').ids.P1.text
        self.P2 = self.root.get_screen('tela').ids.P2.text
        resultados = Workbook()
                
        with open(self.P1) as f:
            reader = csv.reader(f,delimiter=";")

            for row in reader:
                P1_ativo.append(row)
   
   
        with open(self.P2) as f:
            reader = csv.reader(f,delimiter=";")

            for row in reader:
                P2_ativo.append(row)
                
        df_Grupo = self.criaDados()

        


        have_name = False
        have_contrato = False

        wbp1 = Workbook()
        wap1 = wbp1.worksheets[0]
        wap1['A1']='Vendedor'
        wap1['B1']='Total comição'
        wap1['C1']='Total franquia'
        wap1['D1']='Número de contratos'
        wap1['E1']='Número de estornos'

        wap1['B2']='0'
        wap1['C2']='0'
        wap1['D2']='0'
        wap1['E2']='0'


        wbp2 = Workbook()
        wap2 = wbp2.worksheets[0]
        wap2['A1']='Vendedor'
        wap2['B1']='Total comição'
        wap2['C1']='Total franquia'
        wap2['D1']='Número de contratos'
        wap2['E1']='Número de estornos'

        wap2['B2']='0'
        wap2['C2']='0'
        wap2['D2']='0'
        wap2['E2']='0'

        wb_contrato = Workbook()
        wb_c_ativo = wb_contrato.active



        #row contrato = 12
        #row vendedor = 14
        #row cidades = 16
        #row comição = 7
        #


        contador = 0
        df_Grupoa = df_Grupo.active

        for row in df_Grupoa:
            have_contrato = False

            have_name = False
            
            if row[0].value == None:
                pass
            
            else:
                contrato_completo = row[1].value
                #print(row[14].value)
                #print(contrato_completo)

                for rows in P1_ativo:

                    #####P1
                    
                    if rows[19].value == contrato_completo and rows[0].value[0:10] == row[2].value:
                        
                        have_contrato = True
                        #print('valor franquia',rows[12].value,contrato_completo, rows[0].value)
                        
                        for nome in wap1:
                            
                            if row[0].value == nome[0].value:
                                have_name = True

                                break
                        
                        
                        if have_name:
                            
                            for linha in wap1:

                                if row[0].value == linha[0].value:
                                    
                                    valor = float (linha[1].value)
                                    
                                    valor +=float (rows[7].value.replace(',','.'))
                                    linha[1].value = valor
                                    
                                    valor = float (linha[2].value)
                                    valor += float (rows[12].value.replace(',','.'))
                                    linha[2].value = valor
                                    linha[3].value += 1
                                    
                                    break
                        else:
                            for linha in wap1:
                                
                                if linha[0].value == None:
                                    linha[0].value = row[0].value
                                    linha[1].value = float(rows[7].value.replace(',','.'))
                                    linha[2].value = float(rows[12].value.replace(',','.'))
                                    linha[3].value = 1
                                    wap1.insert_rows(2)
                                    
                                    break
                        break
                    ####P2
                have_name = False
                for rows in P2_ativo:
                    
                    print(rows[0].value[0:10] ,'             ' ,row[2].value)
                    if rows[19].value == contrato_completo and rows[0].value[0:10] == row[2].value:
                        
                        
                        have_contrato = True
                        #print('valor franquia',rows[12].value,contrato_completo, rows[0].value)
                        
                        for nome in wap2:
                            
                            if row[0].value == nome[0].value:
                                have_name = True

                                break
                        
                        
                        if have_name:
                            
                            for linha in wap2:

                                if row[0].value == linha[0].value:
                                    
                                    valor = float (linha[1].value)
                                    
                                    valor +=float (rows[7].value.replace(',','.'))
                                    linha[1].value = valor
                                    
                                    valor = float (linha[2].value)
                                    valor += float (rows[12].value.replace(',','.'))
                                    linha[2].value = valor
                                    linha[3].value += 1
                                
                                    break
                        else:
                            for linha in wap2:
                                
                                if linha[0].value == None:
                                    linha[0].value = row[0].value
                                    linha[1].value = float(rows[7].value.replace(',','.'))
                                    linha[2].value = float(rows[12].value.replace(',','.'))
                                    linha[3].value = 1
                                    wap2.insert_rows(2)
                                    
                                    
                                    
                                    break
                        break
                        
                if have_contrato:
                    
                    contador+=1
                
                else:
                    ultimalinha = len(wb_c_ativo['A'])+1
                    wb_c_ativo[f'A{ultimalinha}']=row[0].value
                    wb_c_ativo[f'B{ultimalinha}']=contrato_completo
                    wb_c_ativo[f'C{ultimalinha}']=row[2].value
                    
                            
                            
                    
                
        print(contador)
        df_estornos = Workbook()



        estorno = df_estornos.active

        estorno['A1']='Vendedor'
        estorno['B1']='Contrato'
        estorno['C1']='Data'
        estorno['D1']='Comição'
        estorno['E1']='Comição com fator'
        estorno['F1']='Fator'
        estorno['G1']='P'



        df_banco = load_workbook('banco_contratos.xlsx')
        banco = df_banco.active

        for linha_P in P1_ativo:
            if linha_P[2].value[0:7] == 'ESTORNO':
                
                #print(linha_P[19].value)
                for linha_banco in banco:

                    if linha_banco[1].value == linha_P[19].value and  linha_P[0].value[0:10]== linha_banco[2].value :
                        a = len(estorno['A'])+1

                        vendedor = linha_banco[0].value
                        contrato = linha_banco[1].value
                        data = linha_P[0].value[0:10]     
                        comicao = float (linha_P[7].value.replace(',','.'))
                        fator = float(linha_P[23].value.replace(',','.'))
                        comicao_com_fator = comicao / fator
                        
                        
                        estorno[f'A{a}']=vendedor
                        estorno[f'B{a}']=contrato
                        estorno[f'C{a}']=data
                        estorno[f'D{a}']=comicao
                        estorno[f'E{a}']=comicao_com_fator
                        estorno[f'F{a}']=fator

                        estorno[f'G{a}']='P1'

                                



        for linha_P in P2_ativo:
            if linha_P[2].value[0:7] == 'ESTORNO':
                for linha_banco in banco:
                    
                    if linha_banco[1].value == linha_P[19].value and  linha_P[0].value[0:10]== linha_banco[2].value :
                        
                        a = len(estorno['A'])+1

                        vendedor = linha_banco[0].value
                        contrato = linha_banco[1].value
                        data = linha_P[0].value[0:10]     
                        comicao = float (linha_P[7].value.replace(',','.'))
                        fator = float(linha_P[23].value.replace(',','.'))
                        comicao_com_fator = comicao / fator
                        
                        
                        estorno[f'A{a}']=vendedor
                        estorno[f'B{a}']=contrato
                        estorno[f'C{a}']=data
                        estorno[f'D{a}']=comicao
                        estorno[f'E{a}']=comicao_com_fator
                        estorno[f'F{a}']=fator
                        estorno[f'G{a}']='P2'





       
        

        


        df_estornos.save('estornos.xlsx')
        wbp1.save('dados p1.xlsx')
        wbp2.save('dados p2.xlsx')
        wb_contrato.save('contratos não encontrados.xlsx')    
        
    

            
      
          
if __name__ == '__main__':
    claroApp().run()