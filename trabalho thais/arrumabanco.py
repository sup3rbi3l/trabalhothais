import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from unidecode import unidecode



df = load_workbook('banco_contratos.xlsx')


df_ativo = df.worksheets[0]


for row in df_ativo:
    
    if row[2].value != None:
        data = row[2].value
        #print(data)
        dia,mes,ano = data.split('-')
        
        ##print(contrato_completo)
        ##print(vendedor)
        
        if len(ano) == 4:
                
            row[2].value=f'{ano}-{mes}-{dia}'
            
        else:
                
            row[2].value=f'{dia}-{mes}-{ano}' 

df.save('de prima.xlsx')