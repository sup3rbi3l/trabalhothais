import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE 
from openpyxl.styles import NamedStyle
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from openpyxl import formatting, styles, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from unidecode import unidecode






dados_cidades = {
    "ADAMANTINA": '323',
    "ALVORADA SUL": '021',
    "ALVORADO SUL": '021',
    'ALVORADA DO SUL':'021',
    "AGUDOS": '411',
    "ALAGOINHAS": '802',
    "ALEGRETE": '390',
    "ALMIRANTE TAMANDARE": '294',
    "ALUMINIO": '535',
    "ALVARES MACHADO": '551',
    "ALVORADA": '458',
    "AMERICANA": '386',
    "AMERICO BRASILIENSE": '624',
    "AMPARO": '538',
    "ANANINDEUA": '178',
    "ANAPOLIS": '015',
    "ANDRADINA": '697',
    "APARECIDA": '399',
    "APARECIDA DE GOIANIA": '895',
    "APUCARANA": '322',
    "AQUIRAZ": '561',
    "ARACAJU": '550',
    "ARACATUBA": '404',
    "ARACRUZ": '157',
    "ARAGUAINA": '679',
    "ARAGUARI": '455',
    "ARAPIRACA": '214',
    "ARAPONGAS": '093',
    "ARARANGUA": '400',
    "ARARAQUARA": '409',
    "ARARAS": '413',
    "ARAUCARIA": '333',
    "ARAXA": '151',
    "ARIQUEMES": '870',
    "ARMACAO DOS BUZIOS": '116',
    "ARROIO DO MEIO": '618',
    "ARROIO DO MEIA": '618',
    "ARTUR NOGUEIRA": '419',
    "ARUJA": '420',
    "ATIBAIA": '423',
    "AVARE": '092',
    "BADY BASSITT": '124',
    "BAGE": '690',
    "BALNEARIO CAMBORIU": '752',
    "BARBACENA": '470',
    "BARRA MANSA": '100',
    "BARREIRAS": '353',
    "BARRETOS": '302',
    "BARRINHA": '442',
    "BARUERI": '443',
    "BATATAIS": '448',
    "BAURU": '008',
    "BEBEDOURO": '388',
    "BELEM": '194',
    "BELFORD ROXO": '105',
    "BELO HORIZONTE": '013',
    "BENTO GONCALVES": '688',
    "BERTIOGA": '456',
    "BETIM": '514',
    "BIGUACU": '058',
    "BIRIGUI": '450',
    "BLUMENAU": '700',
    "BOITUVA": '464',
    "BOTUCATU": '471',
    "BRAGANCA PAULISTA": '474',
    "BRASILIA": '040',
    "BRUSQUE": '070',
    "CABEDELO": '133',
    "CABO FRIO": '115',
    "CABREUVA": '787',
    "CACADOR": '819',
    "CACAPAVA": '487',
    "CACHOEIRA DO SUL": '394',
    "CACHOEIRA PAULISTA": '488',
    "CACHOEIRINHA": '560',
    "CACHOEIRO DE ITAPEMIRIM": '911',
    "CACOAL": '898',
    "CAIEIRAS": '498',
    "CAJAMAR": '886',
    "CALDAS NOVAS": '073',
    "CAMACARI": '750',
    "CAMAQUA": '466',
    "CAMBE": '379',
    "CAMBORIU": '868',
    "CAMBURIU": '868',
    "CAMPINA GRANDE": '847',
    "CAMPINAS": '052',
    "CAMPO BOM": '584',
    "CAMPO GRANDE": '011',
    "CAMPO LARGO": '389',
    "CAMPO LIMPO PAULISTA": '012',
    "CAMPOS DO JORDAO": '028',
    "CAMPOS DOS GOYTACAZES": '529',
    "CANELA": '601',
    "CANOAS": '603',
    "CAPAO DA CANOA": '087',
    "CAPAO DA CANOS":'087',
    "CAPIVARI": '530',
    "CARAGUATATUBA": '553',
    "CARAPICUIBA": '533',
    "CARAZINHO": '614',
    "CARIACICA": '917',
    "CARLOS BARBOSA": '920',
    "CARUARU": '152',
    "CASA BRANCA": '202',
    "CASCAVEL": '641',
    "CASTANHAL": '023',
    "CATAGUASES": '340',
    "CATANDUVA": '543',
    "CAXIAS DO SUL": '687',
    "CERQUILHO": '296',
    "CHAPECO": '068',
    "CHARQUEADAS": '217',
    "CIANORTE": '430',
    "COLATINA": '489',
    "COLOMBO": '437',
    "CONCORDIA": '113',
    "CONSELHEIRO LAFAIETE": '741',
    "CONTAGEM": '753',
    "CORDEIROPOLIS": '432',
    "CORONEL FABRICIANO": '778',
    "COSMOPOLIS": '565',
    "COTIA": '567',
    "CRAVINHOS": '569',
    "CRICIUMA": '089',
    "CRUZ ALTA": '685',
    "CRUZEIRO": '572',
    "CUBATAO": '573',
    "CUIABA": '719',
    "CURITIBA": '884',
    "DESCALVADO": '576',
    "DIADEMA": '577',
    "DIVINOPOLIS": '832',
    "DOIS IRMAOS": '781',
    "DOURADOS": '867',
    "DRACENA": '586',
    "DUQUE DE CAXIAS": '161',
    "ELDORADO DO SUL": '918',
    "ELIAS FAUSTO": '595',
    "EMBU DAS ARTES": '599',
    "ENCANTADO": '934',
    "ERECHIM": '695',
    "ESPIRITO SANTO DO PINHAL": '970',
    "ESTANCIA VELHA": '757',
    "ESTEIO": '758',
    "ESTRELA": '231',
    "EUNAPOLIS": '061',
    "EUSEBIO": '044',
    "FARROUPILHA": '083',
    "FEIRA DE SANTANA": '852',
    "FERNANDOPOLIS": '098',
    "FLORIANOPOLIS": '088',
    "FLORIPA":'088',
    "FORMOSA": '782',
    "FORTALEZA": '097',
    "FOZ DO IGUACU": '506',
    "FRAIBURGO": '449',
    "FRANCA": '056',
    "FREDERICO WESTPHALEN": '536',
    "GARCA": '232',
    "GARIBALDI": '568',
    "GASPAR": '154',
    "GOIANIA": '010',
    "GOVERNADOR VALADARES": '930',
    "GRAMADO": '817',
    "GRAVATAI": '720',
    "GUAIBA": '728',
    "GUAIRA": '364',
    "GUAPIACU": '381',
    "GUARAMIRIM": '163',
    "GUARAPUAVA": '531',
    "GUARARAPES": '496',
    "GUARATINGUETA": '656',
    "GUARATUBA": '081',
    "GUARUJA": '659',
    "GUARULHOS": '661',
    "GURUPI": '992',
    "HERVAL DO OESTE": '698',
    "HERVAL D OESTE": '698',
    "HORTOLANDIA": '668',
    "IBATE": '674',
    "IBIPORA": '204',
    "IBIUNA": '755',
    "ICARA": '724',
    "IGREJINHA": '885',
    "ILHEUS": '843',
    "IMBE": '835',
    "IMPERATRIZ": '051',
    "INDAIAL": '775',
    "INDAIATUBA": '054',
    "IPATINGA": '405',
    "IPERO": '951',
    "ITABIRA": '813',
    "ITABUNA": '084',
    "ITAGUAI": '057',
    "ITAJAI": '193',
    "ITAJUBA": '402',
    "ITANHAEM": '712',
    "ITAPECERICA DA SERRA": '714',
    "ITAPEMA": '094',
    "ITAPETININGA": '715',
    "ITAPEVA": '166',
    "ITAPEVI": '722',
    "ITAPEIRA": '183',
    "ITAQUAQUECETUBA": '733',
    "ITAQUI": '106',
    "ITATIBA": '739',
    "ITAUNA": '144',
    "ITU": '746',
    "ITUIUTABA": '060',
    "ITUMBIARA": '257',
    "ITUPEVA": '360',
    "ITUVERAVA": '574',
    "IVOTI": '173',
    "JABOATAO DOS GUARARAPES": '223',
    "JABOTICABAL": '410',
    "JACAREI": '756',
    "JAGUARIUNA": '761',
    "JALES": '492',
    "JANDIRA": '765',
    "JARAGUA DO SUL": '203',
    "JARDINOPOLIS": '575',
    "JARINU": '581',
    "JATAI": '303',
    "JAU": '769',
    "JEQUIE": '718',
    "JI-PARANA": '994',
    "JOACABA": '119',
    "JOAO MONLEVADE": '526',
    "JOAO PESSOA": '907',
    "JOINVILLE": '086',
    "JOIVILLE": '086',
    "JOSE BONIFACIO": '662',
    "JUAZEIRO": '823',
    "JUAZEIRO DO NORTE": '110',
    "JUIZ DE FORA": '126',
    "JUNDIAI": '055',
    "LAGES": '213',
    "LAGE": '213',
    "LAGOA SANTA": '043',
    "LAJEADO": '689',
    "LARANJAL PAULISTA": '859',
    "LAURO DE FREITAS": '042',
    "LAVRAS": '909',
    "LEME": '905',
    "LENCOIS PAULISTA": '913',
    "LIMEIRA": '791',
    "LINS": '948',
    "LONDRINA": '996',
    "LORENA": '795',
    "LOUVEIRA": '798',
    "LUCAS DO RIO VERDE": '875',
    "MACAE": '197',
    "MACAPA": '393',
    "MACEIO": '480',
    "MAFRA": '348',
    "MAIRINQUE": '810',
    "MANAUS": '121',
    "MANHUACU": '222',
    "MARABA": '002',
    "MARAU": '638',
    "MARECHAL CANDIDO RONDON": '208',
    "MARILIA": '820',
    "MARINGA": '091',
    "MATA DE SAO JOAO": '481',
    "MATAO": '297',
    "MAUA": '833',
    "MEDIANEIRA": '428',
    "MESQUITA": '239',
    "MIGUEL PEREIRA": '391',
    "MIRANDOPOLIS": '392',
    "MIRASSOL": '850',
    "MOCOCA": '431',
    "MOGI DAS CRUZES": '856',
    "MOGI GUACU": '863',
    "MOGI MIRIM": '865',
    "MONGAGUA": '869',
    "MONTE ALTO": '534',
    "MONTE MOR": '882',
    "MONTENEGRO": '950',
    "MONTES CLAROS": '275',
    "MORUNGABA": '626',
    "MOSSORO": '947',
    "NATAL": '095',
    "NAVEGANTES": '260',
    "NILOPOLIS": '226',
    "NITEROI": '228',
    "NOVA FRIBURGO": '030',
    "NOVA IGUACU": '237',
    "NOVA LIMA": '312',
    "NOVA MUTUM": '341',
    "NOVA ODESSA": '910',
    "NOVA PETROPOLIS": '201',
    "NOVO HAMBURGO": '686',
    "OLIMPIA": '987',
    "OLINDA": '269',
    "ORLANDIA": '592',
    "OSASCO": '924',
    "OSORIO": '293',
    "OURINHOS": '071',
    "PAICANDU": '912',
    'PALHOCA': '278',
    'PALMAS': '540',
    'PALMEIRA DAS MISSOES': '412',
    'PANAMBI': '447',
    'PARA DE MINAS': '812',
    'PARAGOMINAS': '628',
    'PARAIBA DO SUL': '677',
    'PARAISO DO TOCANTINS': '085',
    'PARANAGUA': '749',
    'PARAUAPEBAS': '745',
    'PARNAIBA': '069',
    'PARNAMIRIM': '637',
    'PAROBE': '499',
    'PASSO FUNDO': '693',
    'PASSOS': '995',
    'PATOS DE MINAS': '014',
    'PAULINIA': '954',
    'PAULISTA': '291',
    'PEDREIRA': '966',
    'PEDRO LEOPOLDO': '234',
    'PELOTAS': '691',
    'PENAPOLIS': '495',
    'PERUIBE': '972',
    'PETROLINA': '956',
    'PETROPOLIS': '250',
    'PIEDADE': '571',
    'PINDAMONHANGABA': '976',
    'PINHAIS': '770',
    'PIRACAIA': '657',
    'PIRACICABA': '009',
    'PIRAQUARA': '380',
    'PIRASSUNUNGA': '747',
    'POA': '609',
    'POCOS DE CALDAS': '037',
    'PONTA GROSSA': '794',
    'PONTA PORA': '017',
    'PONTAL': '844',
    'PORTO ALEGRE': '078',
    'PORTO FELIZ': '022',
    'PORTO FERREIRA': '908',
    'PORTO SEGURO': '532',
    'PORTO VELHO': '220',
    'POTIM': '024',
    'POTIRENDABA': '941',
    'POUSO ALEGRE': '459',
    'PRAIA GRANDE': '027',
    'PRESIDENTE BERNARDES': '025',
    'PRESIDENTE PRUDENTE': '036',
    'PROMISSAO': '074',
    'QUATRO BARRAS': '683',
    'RAFARD': '048',
    'RECIFE': '136',
    'REGISTRO': '211',
    'RESENDE': '268',
    'RIBEIRAO PIRES': '067',
    'RIBEIRAO PRETO': '005',
    'RIO BRANCO': '066',
    'RIO CLARO': '076',
    'RIO DAS OSTRAS': '285',
    'RIO DE JANEIRO': '038',
    'RIO DO SUL': '301',
    'RIO GRANDE': '692',
    'RIO NEGRINHO': '352',
    'RIO PARDO': '478',
    'RIO VERDE': '112',
    'ROLANDIA': '851',
    'RONDONOPOLIS': '818',
    'ROSARIO DO SUL': '613',
    'SABARA': '541',
    'SALTO': '295',
    'SALVADOR': '230',
    'SANTA BARBARA DO OESTE': '103',
    'SANTA CRUZ DO RIO PARDO': '108',
    'SANTA CRUZ DO SUL': '694',
    'SANTA CRUZ': '694',
    'SANTA GERTRUDES': '793',
    'SANTA HELENA DE GOIAS': '580',
    'SANTA ISABEL': '808',
    'SANTA LUZIA': '072',
    'SANTA MARIA': '075',
    'SANTA ROSA': '940',
    'SANTA ROSA DE VITERBO': '904',
    'SANTANA': '616',
    'SANTANA DE PARNAIBA': '621',
    'SANT ANA DO LIVRAMENTO': '041',
    'SANTO ANDRE': '129',
    'SANTO ANGELO': '077',
    'SANTOS': '004',
    'SAO BENTO': '570',
    'SAO BENTO ': '570',
    'SAO BERNARDO DO CAMPO': '141',
    'SAO BORJA': '229',
    'SAO CAETANO DO SUL': '143',
    'SAO CARLOS': '053',
    'SAO FRANCISCO DO SUL': '620',
    'SAO GABRIEL': '270',
    'SAO GONCALO': '305',
    'SAO JOAO DA BOA VISTA': '200',
    'SAO JOAO DE MERITI': '316',
    'SAO JOAO DEL REI': '273',
    'SAO JOAQUIM DA BARRA': '615',
    'SAO JOSE': '363',
    'SAO JOSE DO RIO PARDO': '315',
    'SAO JOSE DO RIO PRETO': '006',
    'SAO JOSE DOS CAMPOS': '162',
    'SAO JOSE DOS PINHAIS': '923',
    'SAO LEOPOLDO': '710',
    'SAO LOURENCO DO SUL': '482',
    'SAO LUIS': '096',
    'SAO LUIZ GONZAGA': '504',
    'SAO PAULO': '003',
    'SAO PEDRO DA ALDEIA': '320',
    'SAO ROQUE': '678',
    'SAO SEBASTIAO': '182',
    'SAO VICENTE': '187',
    'SAPIRANGA': '330',
    'SAPUCAIA DO SUL': '332',
    'SENADOR CANEDO': '866',
    'SERRA': '566',
    'SERRA NEGRA': '579',
    'SERRANA': '587',
    'SERTAOZINHO': '195',
    'SETE LAGOAS': '742',
    'SINOP': '837',
    'SOBRAL': '521',
    'SOROCABA': '007',
    'SORRISO': '838',
    'SUMARE': '207',
    'SUZANO': '209',
    'TABOAO DA SERRA': '216',
    'TAMBAU': '854',
    'TAQUARA': '242',
    'TATUI': '233',
    'TAUBATE': '235',
    'TEIXEIRA DE FREITAS': '807',
    'TELEMACO BORBA': '331',
    'TELEMACO': '331',
    'TEOFILO OTONI': '763',
    'TERESINA': '063',
    'TERESOPOLIS': '338',
    'TEUTONIA': '324',
    'TIETE': '241',
    'TIMBO': '026',
    'TIMON': '165',
    'TIMOTEO': '477',
    'TORRES': '398',
    'TRAMANDAI': '438',
    'TRAMADAI': '438',
    'TREMEMBE': '246',
    'TRES CORACOES': '780',
    'TRES COROAS': '497',
    'TRES LAGOAS': '378',
    'TRES RIOS': '771',
    'TRINDADE': '047',
    'TUBARAO': '090',
    'UBA': '164',
    'UBATUBA': '261',
    'UBERABA': '803',
    'UBERLANDIA': '806',
    'UNIAO DA VITORIA': '593',
    'UNIAO DA V': '593',
    'UNIAO': '593',
    'UNIAO ': '593',
    'URUGUAIANA': '684',
    'VACARIA': '764',
    'VALENCA': '800',
    'VALINHOS': '272',
    'VALPARAISO': '362',
    'VALPARAISO DE GOIAS': '160',
    'VARGEM GRANDE PAULISTA': '276',
    'VARGINHA': '821',
    'VARZEA GRANDE': '858',
    'VARZEA PAULISTA': '277',
    'VASSOURAS': '836',
    'VENANCIO AIRES': '829',
    'VENANCIO': '829',
    'VENANCIO ': '829',
    'VERA CRUZ': '839',
    'VERANOPOLIS': '853',
    'VESPASIANO': '834',
    'VIAMAO': '871',
    'VICOSA': '588',
    'VIDEIRA': '236',
    'VILA VELHA': '507',
    'VILHENA': '206',
    'VINHEDO': '279',
    'VITORIA': '508',
    'VITORIA DA CONQUISTA': '336',
    'VOLTA REDONDA': '359',
    'VOTORANTIM': '282',
    'VOTUPORANGA': '559',
    'XANXERE': '334',
    'XAXIM': '361',
    'JARAGUA': '203',
    'SAO BENTO DO SUL':'000',
    'PORTO UNIAO':'000'
}
 
wbT = Workbook()
waT = wbT.active
        
waT['A1']='Data'
waT['B1']='Contrato'
waT['C1']='Vendedor'


df_Grupo = load_workbook('grupo.xlsx', keep_vba=True, data_only=True)
gian = load_workbook('gian.xlsx', keep_vba=True, data_only=True)
ag=gian.worksheets[0]
fusao = Workbook()
f_ativo = fusao.active


#print(df_Grupo.sheetnames)
tabela = 0 
for grupo in df_Grupo.worksheets:

    
    for i in range(1,26):
        df = df_Grupo.worksheets[tabela]
        #print (df.cell(row=1,column=i).value)
        
        if df.cell(row=1,column=i).value == 'Dia inst':
            data_instalacao_row=i-1
        elif df.cell(row=1,column=i).value == 'Contrato':
            contrato_row=i-1
        elif df.cell(row=1,column=i).value == 'Instalado?':
            instalado_row=i-1
        elif df.cell(row=1,column=i).value == 'Vendedor':
            vendedor_row=i-1
        elif df.cell(row=1,column=i).value == 'Cidade':
            cidade_row=i-1
        elif df.cell(row=1,column=i).value == 'Cod Cidade':
            cod_cidade_row=i-1
   

    for row in grupo:

        
        if row[0].value == None or grupo.title == 'dados':
                break
        
        elif row[cidade_row].value == 'Cidade' or row[contrato_row].value == '---' or row[contrato_row].value == None or row[instalado_row].value != 'INSTALADA' or row[data_instalacao_row].value == None or (row[cidade_row].value == None and (row[cod_cidade_row].value == None or row[cod_cidade_row].value == '#N/A')) :
            pass    
        
        else:
            
            if row[cod_cidade_row].value != None and row[cod_cidade_row].value != '#N/A':
                cod = str (int (row[cod_cidade_row].value))
                print( row[cod_cidade_row].value)
                print(row[cod_cidade_row].value != '#N/A')
                
                while  len(cod) <3:
                    cod = '0'+cod 
                
                
                contrato_completo = str(cod)+str(row[contrato_row].value)
                print(contrato_completo)
            else:
                
                contrato_completo = str (dados_cidades [unidecode(row[cidade_row].value.upper())])+str(row[contrato_row].value)
                print(dados_cidades [unidecode(row[cidade_row].value.upper())])
                print(row[cidade_row].value)
                print(contrato_completo)
                
            vendedor = row[vendedor_row].value
            data = str(row[data_instalacao_row].value)[0:10].replace('/','-')

            ano,mes,dia = data.split('-')

            a = len(f_ativo['A'])+1
            f_ativo[f'A{a}']=vendedor
            f_ativo[f'B{a}']=contrato_completo
            
            if len(ano) == 4:
                
                f_ativo[f'C{a}']=f'{ano}-{mes}-{dia}'
                
            else:
                
                f_ativo[f'C{a}']=f'{dia}-{mes}-{ano}' 
                
    tabela+= 1
                
                
for row in ag:
    
    for i in range(1,26):
        
        if ag.cell(row=1,column=i).value == 'DATA DA INSTALAÇÃO' or ag.cell(row=1,column=i).value == 'Dia inst':
            data_instalacao_row=i-1
        elif ag.cell(row=1,column=i).value == 'CONTRATO' or ag.cell(row=1,column=i).value == 'Contrato':
            contrato_row=i-1
        elif ag.cell(row=1,column=i).value == 'INSTALADO' or ag.cell(row=1,column=i).value == 'Instalado?':
            instalado_row=i-1
        elif ag.cell(row=1,column=i).value == 'CIDADE' or ag.cell(row=1,column=i).value == 'Cidade':
            cidade_row=i-1

    
    if row[cidade_row].value == 'Cidade' or row[cidade_row].value == '---' or row[cidade_row].value == None or row[cidade_row].value == None or (row[instalado_row].value != 'ok' and row[instalado_row].value != 'INSTALADA') or row[data_instalacao_row].value == None or (row[cidade_row].value == None and (row[cod_cidade_row].value == None or row[cod_cidade_row].value == '#N/A')):

        pass
    elif row[0].value == None:
        break
        
    else:
        
        contrato_completo = str (dados_cidades [unidecode(row[cidade_row].value.upper())])+str(row[contrato_row].value)
        
        
        vendedor = row[vendedor_row].value
        data = str(row[data_instalacao_row].value)[0:10].replace('/','-')
        dia,mes,ano = data.split('-')

        a = len(f_ativo['A'])+1
        f_ativo[f'A{a}']='GIAN'
        f_ativo[f'B{a}']=contrato_completo
        if len(ano) == 4:
                
            f_ativo[f'C{a}']=f'{ano}-{mes}-{    dia}'
                
        else:
                
            f_ativo[f'C{a}']=f'{dia}-{mes}-{ano}' 
        
                
                

                
#data instalação = 10
#contrato = 13
#cidade = 17

                
                
fusao.save('grupo_e_gian.xlsx')

                